import io
import openpyxl
from openpyxl.styles import Alignment # 줄바꿈을 위해 Alignment만 남김
from Terms_Analyze.schemas.MVP_dto import TermsResponse
from Company_Terms_Analzye.schemas.Company_dto import CompanyAnalysisResponse

def create_consumer_report_excel(data: TermsResponse) -> io.BytesIO:
    wb = openpyxl.Workbook() # 엑셀 워크북 생성
    
    ws_summary = wb.active # 1번 시트 - 요약
    ws_summary.title = "불공정 약관 요약"
    
    # 열 제목 지정 및 데이터 입력
    ws_summary.append(["항목", "내용"])
    ws_summary.append(["보고서 제목", data.summary.title])
    ws_summary.append(["개요", data.summary.overview])
    ws_summary.append(["총 조항 수", data.summary.totalClauses])
    ws_summary.append(["불공정 조항 수", data.summary.unfairCount])
    ws_summary.append(["위험도", data.summary.riskLevel])
    
    # 2번 시트 - 불공정 조항 상세
    ws_detail = wb.create_sheet("불공정 조항 상세")
    ws_detail.append(["ID", "조항 번호", "약관 원문", "이슈 유형", "설명", "심각도", "관련 법률"])
    
    for clause in data.unfairClauses:
        for issue in clause.issues:
            ws_detail.append([
                clause.id,
                clause.clauseNumber,
                clause.text,
                issue.type,
                issue.description,
                issue.severity,
                issue.relatedLaw
            ])
    
    # 2번 시트 각 셀마다 설정 ( 텍스트 줄바꿈 기능 ON )
    for row in ws_detail.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def create_company_report_excel(data: CompanyAnalysisResponse) -> io.BytesIO:
    wb = openpyxl.Workbook() # 워크북 생성
    
    ws_summary = wb.active
    ws_summary.title = "기업용 약관 취약점 분석 요약"  # 1번 시트
    
    ws_summary.append(["항목", "내용"])
    ws_summary.append(["종합 위험도", data.riskLevel])
    ws_summary.append(["요약 보고", data.summary])
    ws_summary.append(["최악의 시나리오", data.worstScenario])
    
    # 2번 시트 - 취약 조항 상세
    ws_detail = wb.create_sheet("취약 조항 상세")
    ws_detail.append(["조항 원문", "위험도", "유형", "관련 법률", "설명", "수정 제안"])
    
    for vuln in data.vulnerabilities:
        ws_detail.append([
            vuln.clause,
            vuln.riskLevel,
            vuln.vulnerabilityType,
            vuln.relatedLaw,    
            vuln.description,
            vuln.suggestion    
        ])
        
    for row in ws_summary.iter_rows(): # 1번 시트 줄바꿈 설정
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws_detail.iter_rows(): # 2번 시트 줄바꿈 설정
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer